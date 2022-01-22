'''
This script pulls the coordinates (UTME, UTMN) from the SnowEx pit forms and
writes the info with header info to a csv.

Intended for quick plotting of UTMs as a check on conversion and transcriptions.

This script works for the __2021__SnowEx__Pit__Form__

This script was modified from the pits_xls2csv_v6.py script from C. Vuyovich.

'''

__author__ = "Megan Mason, ATAAerospace Data Management Liaison"
__version__ = "01"
__maintainer__ = "Megan Mason"
__email__ = "meganmason491@boisestate.edu"
__status__ = "Dvp"
__date__ = "09.2021"

import datetime
import glob
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from pyproj import Proj, transform
from csv import writer
import textwrap
import utm

from zone_dict import dict1

#----------------------------------METHODS--------------------------------------
def readSnowpit(filename):

        # open excel file
        wb = load_workbook(filename)
        ws = wb.active

        # date / location / stie / pit-ID / coordinates
        date = (ws['H3'].value).date()

        location = ws['B3'].value
        site = ws['B6'].value
        pitID = (ws['B8'].value).split('_')[0]

        lat = ws['J8'].value
        lon = ws['N8'].value


    # lat/lon corrections-------------------------------------------------------

        # empty??
        if lat == None:
            LAT = np.nan
            LON = np.nan

        else:

            # remove strings
            if type(lat) is str:
                lat = lat.split('.')[0] # cuts off anything after a decimal (otherwise it would join together and incr. mag. of num)
                lat = float(''.join([i for i in list(lat) if i.isdigit()]))
            if type(lon) is str:
                lon = lon.split('.')[0]
                lon = float(''.join([i for i in list(lon) if i.isdigit()]))

            # If lat = negative, swap lat/lon
            if lat < 0: #recorded as longitude (negative value)
                lat, lon = lon, lat # swap coards

            # convert UTMs to lat/lon
            if lat > 90: #recorded as UTMs
                UTME = lat
                UTMN = lon

                # if UTMN < UTME:
                #     UTMN, UTME = UTME, UTMN # swap coords

                UTMzone = 13
                lat = utm.to_latlon(UTME, UTMN, UTMzone, "Northern")[0] #tuple output, save first

            if lon > 0:
                lon = utm.to_latlon(UTME, UTMN, UTMzone, "Northern")[1] #tuple output, save second


            LAT = round(lat, 5)
            LON = round(lon, 5)


        return date, location, site, pitID, LAT, LON  #lat, lon = NO corrections, #LAT, LON = with corrections

#----------------------------------BODY--------------------------------------

if __name__ == "__main__":

    # set-up
    path_in = Path('/Users/meganmason491/Google Drive/SnowEx-2021/SnowEx-2021-campaign-core/SnowEx2021_TimeSeries_Pits/pre-format-clean/Central-Ag/')

    # data
    data = []
    column_lst = ['Date', 'Location', 'Site', 'PitID', 'latitude', 'longitude']

    for filename in path_in.rglob('*.xlsx'):

        # filename
        print(f"...reading {filename.name}")

        data.append(readSnowpit(filename))

    df = pd.DataFrame(data, columns=column_lst)


    print('No. of data rows:', len(df.index))

    df.to_csv('/Users/meganmason491/Documents/snowex/2021/output/coords.csv', sep=',', header=True)
