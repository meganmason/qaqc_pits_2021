import glob
import os
import shutil
from pathlib import Path

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook

'''
This script is intended to clean up formating issues that were not caught on the 2021 snow pit template.
This clean up will assist the qa/qc process.

    For example:
        - text in some cells is center aligns
        - column widths were adjusted for readablility
        - some text is 'shrink to fit' within a cell (NOTE - this doesn't show up on google slides, but it does once downloaded)

This script is also intended to cover any "name" or "site" changes

    For example:
        - Banner Summit - Snotel --> Banner Snotel (to match with SNEX 2020)
        - ideally give CARC (montana site) a better Location Name.
'''

#------------------------------------------------------------------------------#
# SET-UP
# fname = '/Users/meganmason491/Downloads/IDBRBS_20210115_SNOW_PIT.xlsx'
path_in = Path('/Users/meganmason491/Google Drive/SnowEx-2021/SnowEx-2021-campaign-core/SnowEx2021_TimeSeries_Pits/pre-format')

for filename in path_in.rglob('*.xlsx'):
    print(f"...processing {filename.name}")
    wb = load_workbook(filename)
    ws = wb.active

#------------------------------------------------------------------------------#
# GLOBAL CHANGES
    # Snow Depth (cm) to HS (cm)
    ws['F7:G7'][0][0].value = 'HS\n(cm)'

    # Observers alignment
    ws['J3'].alignment = Alignment(horizontal = "left", vertical = "top")

    # Latitude alignment
    ws['J8'].alignment = Alignment(horizontal = "center")

    # temp start/end - center
    ws['U6'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['V6'].alignment = Alignment(horizontal = "center", vertical = "center")

    # gps device - shrink to fit
    ws['U8'].alignment = Alignment(shrink_to_fit=True, wrap_text=True)

    # Comments - wrap text
    ws['W3'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True)

    # reduce divider cell (column l)
    ws.column_dimensions['L'].width = 3

    # enviornment/ground condition headers - shrink to fit
    ws['K43'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True, shrink_to_fit=True)
    ws['K44'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True, shrink_to_fit=True)
    ws['K45'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True, shrink_to_fit=True)
    ws['K46'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True, shrink_to_fit=True)
    ws['K47'].alignment = Alignment(horizontal = "center", vertical = "top", wrap_text=True, shrink_to_fit=True)

    # veg bool - shrink to fit
    ws.column_dimensions['M'].width = 7
    ws.column_dimensions['O'].width = 7
    ws.column_dimensions['Q'].width = 7
    ws.column_dimensions['S'].width = 7

    ws.column_dimensions['N'].width = 7
    ws.column_dimensions['P'].width = 7
    ws.column_dimensions['R'].width = 7
    ws.column_dimensions['T'].width = 7

    # tree canopy - bold
    ws['K48:L48'][0][0].font = Font(size=11, bold=True)

    # save
    wb.save(filename)
#
#
print(f"Part 1 pre-formatting complete")
#------------------------------------------------------------------------------#
# INDIVIDUAL SITE CHANGES
path = Path('/Users/meganmason491/Google Drive/SnowEx-2021/SnowEx-2021-campaign-core/SnowEx2021_TimeSeries_Pits/pre-format')

# set up:
# ws['B3:G3'][0][0].value = Location
# ws['B6:G6'][0][0].value = Site
# ws['B8:E8'][0][0].value = Unique PitID


# boise river
for filename in path.rglob('*IDBRBS*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Banner Snotel'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)

for filename in path.rglob('*IDBRBL*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Bogus Lower'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)

for filename in path.rglob('*IDBRBT*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Bogus Lower Trees'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)

for filename in path.rglob('*IDBRK2*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Mores Creek'
    update_pitID = filename.name[:15]
    ws['B6:G6'][0][0].value = update_site
    ws['B8:E8'][0][0].value = update_pitID
    wb.save(filename)


# central ag
for filename in path.rglob('*MTCA*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_location = 'Central Ag Research Center'
    ws['B3:G3'][0][0].value = update_location
    wb.save(filename)

for filename in path.rglob('*MTCASX*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'SnowEx-1'
    update_pitID = filename.name[:15]
    ws['B6:G6'][0][0].value = update_site
    ws['B8:E8'][0][0].value = update_pitID
    wb.save(filename)

for filename in path.rglob('*MTCAWX*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Wx'
    update_pitID = filename.name[:15]
    ws['B6:G6'][0][0].value = update_site
    ws['B8:E8'][0][0].value = update_pitID
    wb.save(filename)

for filename in path.rglob('*MTCAWH*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Wheat'
    update_pitID = filename.name[:15]
    ws['B6:G6'][0][0].value = update_site
    ws['B8:E8'][0][0].value = update_pitID
    wb.save(filename)


# little cottonwood
for filename in path.rglob('*UTLCAC*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Alta Collins'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)

for filename in path.rglob('*UTLCAW*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Atwater'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)


# senator beck
for filename in path.rglob('*COSBSA*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_site = 'Swamp Angel'
    ws['B6:G6'][0][0].value = update_site
    wb.save(filename)

# fraser
for filename in path.rglob('*COFE*'):
    print(f"processing...{filename.stem}")
    wb = load_workbook(filename)
    ws = wb.active
    update_pitID = filename.name[:15]
    ws['B8:E8'][0][0].value = update_pitID
    wb.save(filename)

print(f"Part 2 pre-formatting complete")
