from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook



# LWC serial # format
if ws['H8:I8'][0][0].value != None:
    lwc_serial = int(ws['H8:I8'][0][0].value)
    ws['H8:I8'][0][0].value = 'WIS00{}A'.format(lwc_serial)
